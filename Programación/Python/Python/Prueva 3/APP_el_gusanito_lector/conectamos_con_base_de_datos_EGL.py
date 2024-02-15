import mysql.connector
import openpyxl
from existe_pagina import *

# Configuración de la conexión a la base de datos
def conectar_bbdd ():

    configuracion = {
    "user": "root",
    "password": "1234",
    "host": "localhost", # Puede ser 'localhost' si la base de datos está en tu máquina
    "database": "filo_egl",
    }

    # Conectar a la base de datos
    conection = mysql.connector.connect(**configuracion)

    return conection

def cerrar_conection (conection):
    conection.commit()
    conection.close()


def insertar_datos ():
    #obtenemos los datos del scraping
    todos_libros = extracxion_de_datos()
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()

    scrip = "insert into libro (portada, nombre, autores, precio, ISBN, tipo) values (%s, %s, %s, %s, %s, %s)"
    # Diccionario que queremos insertar en la tabla
    for i in todos_libros:
        cursor.execute(scrip,(i["portada"],i["nombre"],i["autor"],i["precio"],i["ISBN"],i["tipo"]))

    cerrar_conection(conexion)

#insertar_datos()

def insertar_1_dato (nuevo_libro):
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()

    scrip = "insert into libro (portada, nombre, autores, precio, ISBN, tipo) values (%s, %s, %s, %s, %s, %s)"

    cursor.execute(scrip,(nuevo_libro["portada"],nuevo_libro["nombre"],nuevo_libro["autor"],nuevo_libro["precio"],nuevo_libro["ISBN"],nuevo_libro["tipo"]))

    cerrar_conection(conexion)

#insertar_1_dato({'portada': 'https://www.elgusanitolector.com/imagenes/9788412/978841271541.webp', 'nombre': 'NAVEGAR NUESTRAS GEOGRAFÍAS', 'autor': 'Zibechi, raúl, ', 'precio': 13.5, 'ISBN': '978-84-127154-1-5', 'tipo': 'Libros singulares'})

def borrar_dato_id (id):
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()
    #código
    script = "DELETE FROM libro WHERE id=" + str(id)

    #ejecutamos
    cursor.execute(script)

    cerrar_conection(conexion)

#borrar_dato_id(31)

def extraer_datos():

    #creamos lista
    lista_libro = []

    #abrir conexion
    conexion = conectar_bbdd()

    #abrimos cursor
    cursor = conexion.cursor(dictionary=True)

    #montamos el script
    script = "select * from libro"

    #ejecutamos la consulta
    cursor.execute(script)

    #Obtenemos columnas
    columnas = cursor.column_names

    #obtener los datos de la consulta
    lista_libro = cursor.fetchall()

    #cerramos conexion
    cerrar_conection(conexion)

    return lista_libro,columnas

#extraer_datos()

def crear_excel_con_libros():

    #Datos
    lista_libros,nombres_columnas = extraer_datos()

    #Creamos Hoja excel
    documento = openpyxl.Workbook()
    hoja_1 = documento.active

    #rellenamos la primera fila con los titulos
    for contador, nombre_columna in enumerate(nombres_columnas, start=1):
        hoja_1.cell(row=1,column=contador,value=nombre_columna)


    for fila,libro in enumerate(lista_libros, start=2):
        for columna,campo_libro in enumerate(libro.values(),start=1):
            hoja_1.cell(row=fila, column=columna,value=campo_libro)


    documento.save("libreria_egl_filo.xlsx")

#crear_excel_con_libros()

def buscar_libro_id (id):
    #abrir conexion
    conexion = conectar_bbdd()

    #abrimos cursor
    cursor = conexion.cursor(dictionary=True)

    #montamos el script
    script = "SELECT * FROM libro WHERE id=" + str(id)

    #ejecutamos la consulta
    cursor.execute(script)

    libro = cursor.fetchall()

    return libro[0]
#print(buscar_libro_id(250))

def buscar_tipo (tipo):
    #abrir conexion
    conexion = conectar_bbdd()

    #abrimos cursor
    cursor = conexion.cursor(dictionary=True)

    tipo = "'"+tipo+"'"

    #montamos el script
    script = "select * from libro where tipo = "+ tipo

    #ejecutamos la consulta
    cursor.execute(script)

    libro = cursor.fetchall()

    return libro

#print(buscar_tipo('libros de viaje'))