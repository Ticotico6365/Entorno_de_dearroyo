import mysql.connector
import openpyxl
from el_gusanito_lector_scraping import *

# Configuración de la conexión a la base de datos
def conectar_bbdd ():

    configuracion = {
    "user": "root",
    "password": "1daw",
    "host": "localhost", # Puede ser 'localhost' si la base de datos está en tu máquina
    "database": "filo_egl_scraping",
    }

    # Conectar a la base de datos
    conection = mysql.connector.connect(**configuracion)

    return conection

def cerrar_conection (conection):
    conection.commit()
    conection.close()


def insertar_datos ():
    #obtenemos los datos del scraping
    todos_libros = extracxion_de_datos()
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()

    scrip = "insert into libro (portada, nombre, precio, autores, descripcion, ISBN) values (%s, %s, %s, %s, %s, %s)"
    # Diccionario que queremos insertar en la tabla
    for i in todos_libros:
        cursor.execute(scrip,(i["portada"],i["nombre"],i["precio"],i["autor"],i["descripcion"],i["ISBN"]))

    cerrar_conection(conexion)

#insertar_datos()

def insertar_1_dato (portada, nombre, autor, precio, descripcion, ISBN):
    nuevo_libro = {
        "portada": portada,
        "nombre": nombre,
        "autor": autor,
        "precio": precio,
        "descripcion": descripcion,
        "ISBN": ISBN
    }
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()

    scrip = "insert into libro (portada, nombre, precio, autores, descripcion, ISBN) values (%s, %s, %s, %s, %s, %s)"

    cursor.execute(scrip,(nuevo_libro["portada"],nuevo_libro["nombre"],nuevo_libro["precio"],nuevo_libro["autor"],nuevo_libro["descripcion"],nuevo_libro["ISBN"]))

    cerrar_conection(conexion)

#insertar_1_dato("portada", "nombre", "pepe", 8, "pepe fue a pescar", "987-521-456")

def borrar_dato_id (id):
    #conectamos con la bbdd
    conexion = conectar_bbdd()
    #creamos el cursor
    cursor = conexion.cursor()
    #código
    script = "DELETE FROM libro WHERE id=" + str(id)

    #ejecutamos
    cursor.execute(script)

    cerrar_conection(conexion)

#borrar_dato_id(31)

def extraer_datos():

    #creamos lista
    lista_libro = []

    #abrir conexion
    conexion = conectar_bbdd()

    #abrimos cursor
    cursor = conexion.cursor(dictionary=True)

    #montamos el script
    script = "select * from libro"

    #ejecutamos la consulta
    cursor.execute(script)

    #Obtenemos columnas
    columnas = cursor.column_names

    #obtener los datos de la consulta
    lista_libro = cursor.fetchall()

    #cerramos conexion
    cerrar_conection(conexion)

    return lista_libro,columnas

#extraer_datos()

def crear_excel_con_libros():

    #Datos
    lista_libros,nombres_columnas = extraer_datos()

    #Creamos Hoja excel
    documento = openpyxl.Workbook()
    hoja_1 = documento.active

    #rellenamos la primera fila con los titulos
    for contador, nombre_columna in enumerate(nombres_columnas, start=1):
        hoja_1.cell(row=1,column=contador,value=nombre_columna)


    for fila,libro in enumerate(lista_libros, start=2):
        for columna,campo_libro in enumerate(libro.values(),start=1):
            hoja_1.cell(row=fila, column=columna,value=campo_libro)


    documento.save("libreria_egl_filo.xlsx")

crear_excel_con_libros()